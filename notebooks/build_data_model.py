# build_data_model.py
# تشغيل: python build_data_model.py
# الناتج: churn_data_model.xlsx

import random, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42); np.random.seed(42)
wb = Workbook()

# ── helpers ──────────────────────────────────────────────
def F(bold=False,sz=10,color="1A1A2E",italic=False):
    return Font(name="Arial",bold=bold,size=sz,color=color,italic=italic)
def FILL(h): return PatternFill("solid",fgColor=h)
def CTR(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
def LFT(): return Alignment(horizontal="left",vertical="center")
def RGT(): return Alignment(horizontal="right",vertical="center")
def BDR():
    s=Side(style="thin",color="D1D5DB")
    return Border(left=s,right=s,top=s,bottom=s)
def hrow(ws,row,nc,bg="111118",fg="FFFFFF"):
    for c in range(1,nc+1):
        x=ws.cell(row=row,column=c)
        x.font=F(bold=True,sz=10,color=fg); x.fill=FILL(bg)
        x.alignment=CTR(); x.border=BDR()
    ws.row_dimensions[row].height=22
def dc(ws,row,col,val,fmt=None,bold=False,col_=None,bg=None,ac="left"):
    c=ws.cell(row=row,column=col,value=val)
    c.font=F(bold=bold,sz=10,color=col_ or "1A1A2E")
    c.alignment=CTR() if ac=="center" else RGT() if ac=="right" else LFT()
    c.border=BDR()
    if fmt: c.number_format=fmt
    c.fill=FILL(bg) if bg else (FILL("F9FAFB") if row%2==0 else FILL("FFFFFF"))
    return c


def setw(ws,ws_):
    for i,w in enumerate(ws_,1): ws.column_dimensions[get_column_letter(i)].width=w

# ── generate 2000 rows ───────────────────────────────────
CONTRACTS=["Month-to-month","One year","Two year"]
INTERNET=["Fiber optic","DSL","No internet"]
PAYMENT=["Electronic check","Mailed check","Bank transfer","Credit card"]
rows=[]
for i in range(1,2001):
    cid=f"CUS{i:04d}"; g=random.choice(["Male","Female"])
    sr=random.choices([0,1],weights=[.84,.16])[0]
    pt=random.choice(["Yes","No"]); dp=random.choice(["Yes","No"])
    tn=min(int(np.random.exponential(28))+1,72)
    ph=random.choices(["Yes","No"],weights=[.9,.1])[0]
    ml=random.choice(["Yes","No"]) if ph=="Yes" else "No phone service"
    inet=random.choices(INTERNET,weights=[.44,.34,.22])[0]
    sv=lambda: random.choice(["Yes","No"]) if inet!="No internet" else "No internet service"
    se,bk,pr,tc,st,sm=sv(),sv(),sv(),sv(),sv(),sv()
    ct=random.choices(CONTRACTS,weights=[.55,.25,.20])[0]
    pb=random.choice(["Yes","No"])
    pm=random.choices(PAYMENT,weights=[.34,.23,.22,.21])[0]
    base=20
    if ph=="Yes": base+=20
    if ml=="Yes": base+=10
    if inet=="Fiber optic": base+=40
    elif inet=="DSL": base+=25
    for s in [se,bk,pr,tc,st,sm]:
        if s=="Yes": base+=8
    mo=round(base+random.uniform(-4,9),2)
    tot=round(mo*tn*random.uniform(.93,1.02),2)
    p=min(max(0.08+(0.30 if ct=="Month-to-month" else 0.08 if ct=="One year" else 0)
              +(0.22 if tn<6 else 0.12 if tn<12 else -0.08 if tn>48 else 0)
              +(0.10 if mo>75 else 0)+(0.10 if inet=="Fiber optic" else 0),.03),.93)
    p=round(p,4)
    ch="Yes" if random.random()<p else "No"
    rk="High" if p>.60 else("Medium" if p>.35 else "Low")
    tg="0-12 mo" if tn<=12 else("13-24 mo" if tn<=24 else("25-48 mo" if tn<=48 else "49-72 mo"))
    cg="Low <$35" if mo<35 else("Mid $35-65" if mo<65 else("High $65-90" if mo<90 else "Premium >$90"))
    rows.append([cid,g,sr,pt,dp,tn,ph,ml,inet,se,bk,pr,tc,st,sm,ct,pb,pm,mo,tot,ch,p,rk,tg,cg])

# ═════════════════════════════════════════════════════════
# 00_DATA_MODEL
# ═════════════════════════════════════════════════════════
ws0=wb.active; ws0.title="00_DATA_MODEL"
# title(ws0,"DATA MODEL MAP — STAR SCHEMA",span=6)
# note(ws0,"fact_Customers is the center. All dim_ tables join to it via matching column names.",span=6)
schema=[["TABLE","TYPE","KEY","JOINS_TO","ROWS","PURPOSE"],
["fact_Customers","FACT","CustomerID","All dim_ tables","2,000","Core — one row per customer"],
["dim_Contract","DIM","ContractKey","fact_Customers.Contract","3","Contract types + live churn formulas"],
["dim_Internet","DIM","InternetKey","fact_Customers.InternetService","3","Internet service analysis"],
["dim_Payment","DIM","PaymentKey","fact_Customers.PaymentMethod","4","Payment + auto-pay flag"],
["dim_RiskLevel","DIM","RiskKey","fact_Customers.RiskLevel","3","Risk bands + action plan"],
["dim_TenureGroup","DIM","TenureKey","fact_Customers.TenureGroup","4","Cohort buckets"],
["dim_ChargesGroup","DIM","ChargesKey","fact_Customers.ChargesGroup","4","Charges bands"],
["agg_ChurnSummary","AGG","—","Derived from fact","17","All KPIs auto-calculated"],
["agg_Financial","AGG","—","Derived from fact","10+","Revenue + ROI scenarios"]]
r=3; hrow(ws0,r,6,bg="0A0A0F")
for ci,h in enumerate(schema[0],1): ws0.cell(row=r,column=ci,value=h)
r+=1
tm={"FACT":("1E40AF","DBEAFE"),"DIM":("065F46","D1FAE5"),"AGG":("7C3AED","EDE9FE")}
for rd in schema[1:]:
    for ci,v in enumerate(rd,1):
        c=dc(ws0,r,ci,v)
        if ci==2:
            tc,bc=tm.get(v,("374151","F3F4F6"))
            c.fill=FILL(bc); c.font=F(bold=True,sz=10,color=tc); c.alignment=CTR()
    ws0.row_dimensions[r].height=18; r+=1
r+=1
hrow(ws0,r,4,bg="374151")
for ci,h in enumerate(["FROM (fact column)","","TO (dim key)","TYPE"],1): ws0.cell(row=r,column=ci,value=h)
r+=1
for a,b,c2,d in [["fact_Customers.Contract","→","dim_Contract.ContractKey","Many:1"],
["fact_Customers.InternetService","→","dim_Internet.InternetKey","Many:1"],
["fact_Customers.PaymentMethod","→","dim_Payment.PaymentKey","Many:1"],
["fact_Customers.RiskLevel","→","dim_RiskLevel.RiskKey","Many:1"],
["fact_Customers.TenureGroup","→","dim_TenureGroup.TenureKey","Many:1"],
["fact_Customers.ChargesGroup","→","dim_ChargesGroup.ChargesKey","Many:1"]]:
    c1=dc(ws0,r,1,a); c_=dc(ws0,r,2,b); c_.font=F(bold=True,sz=14,color="00C87A"); c_.alignment=CTR()
    dc(ws0,r,3,c2); dc(ws0,r,4,d,ac="center").font=F(sz=10,color="7C3AED")
    ws0.row_dimensions[r].height=17; r+=1
setw(ws0,[26,6,28,12,8,36]); ws0.freeze_panes="A4"

# ═════════════════════════════════════════════════════════
# fact_Customers
# ═════════════════════════════════════════════════════════
ws1=wb.create_sheet("fact_Customers")
COLS=["CustomerID","Gender","SeniorCitizen","Partner","Dependents","Tenure","PhoneService",
      "MultipleLines","InternetService","OnlineSecurity","OnlineBackup","DeviceProtection",
      "TechSupport","StreamingTV","StreamingMovies","Contract","PaperlessBilling",
      "PaymentMethod","MonthlyCharges","TotalCharges","Churn","ChurnProbability","RiskLevel","TenureGroup","ChargesGroup"]
# title(ws1,"fact_Customers — 2,000 rows × 25 cols | PK: CustomerID",span=25)
# note(ws1,"FACT TABLE. Joins: Contract→dim_Contract | InternetService→dim_Internet | PaymentMethod→dim_Payment | RiskLevel→dim_RiskLevel | TenureGroup→dim_TenureGroup | ChargesGroup→dim_ChargesGroup",span=25)
r=1; hrow(ws1,r,25,bg="0A0A0F")
for ci,h in enumerate(COLS,1): ws1.cell(row=r,column=ci,value=h)
r+=1
for rd in rows:
    ws1.row_dimensions[r].height=15
    for ci,v in enumerate(rd,1):
        fmt='"$"#,##0.00' if ci in(19,20) else("0.0%" if ci==22 else None)
        c=dc(ws1,r,ci,v,fmt=fmt)
        if ci==21:
            if v=="Yes": c.fill=FILL("FEE2E2"); c.font=F(bold=True,sz=10,color="DC2626")
            else:        c.fill=FILL("D1FAE5"); c.font=F(bold=True,sz=10,color="059669")
            c.alignment=CTR()
        if ci==23:
            m={"High":("DC2626","FEE2E2"),"Medium":("92400E","FEF9C3"),"Low":("059669","D1FAE5")}
            tc,bc=m.get(v,("374151","F9FAFB"))
            c.fill=FILL(bc); c.font=F(bold=True,sz=10,color=tc); c.alignment=CTR()
    r+=1
setw(ws1,[10,8,7,8,10,7,7,14,13,14,13,15,12,12,14,17,15,18,13,13,7,14,10,10,12])
ws1.freeze_panes="A4"; print("✓ fact_Customers")

# ═════════════════════════════════════════════════════════
# dim_Contract
# ═════════════════════════════════════════════════════════
ws2=wb.create_sheet("dim_Contract")
# title(ws2,"dim_Contract | Key: ContractKey | JOIN → fact_Customers.Contract",span=8)
# note(ws2,"Churn rates auto-calculated from fact_Customers using COUNTIFS/AVERAGEIF",span=8)
r=1; hrow(ws2,r,8)
for ci,h in enumerate(["ContractKey","ContractName","ChurnRate","AvgTenure_mo","AvgMonthly","RiskCategory","BenchmarkRate","RetentionAction"],1): ws2.cell(row=r,column=ci,value=h)
r+=1
for key,bench,risk,action in [("Month-to-month","42%","HIGH RISK","Offer annual upgrade + 15% discount"),
("One year","11%","MEDIUM RISK","Loyalty reward at renewal"),("Two year","3%","LOW RISK","Maintain quality — ambassador program")]:
    row_r=r
    formulas=[key,key,
    f'=COUNTIFS(fact_Customers!P:P,A{r},fact_Customers!U:U,"Yes")/COUNTIF(fact_Customers!P:P,A{r})',
    f'=AVERAGEIF(fact_Customers!P:P,A{r},fact_Customers!F:F)',
    f'=AVERAGEIF(fact_Customers!P:P,A{r},fact_Customers!S:S)',
    risk,bench,action]
    rcm={"HIGH RISK":("DC2626","FEE2E2"),"MEDIUM RISK":("92400E","FEF9C3"),"LOW RISK":("059669","D1FAE5")}
    for ci,v in enumerate(formulas,1):
        c=dc(ws2,r,ci,v)
        if ci==3: c.number_format="0.0%"; c.alignment=CTR()
        if ci in(4,5): c.number_format='"$"#,##0.00'; c.alignment=RGT()
        if ci==6:
            tc,bc=rcm.get(v,("374151","F9FAFB"))
            c.fill=FILL(bc); c.font=F(bold=True,sz=10,color=tc); c.alignment=CTR()
    ws2.row_dimensions[r].height=18; r+=1
setw(ws2,[18,18,12,13,13,13,13,40]); ws2.freeze_panes="A4"; print("✓ dim_Contract")

# ═════════════════════════════════════════════════════════
# dim_Internet
# ═════════════════════════════════════════════════════════
ws3=wb.create_sheet("dim_Internet")
# title(ws3,"dim_Internet | Key: InternetKey | JOIN → fact_Customers.InternetService",span=7)
# note(ws3,"Fiber optic shows highest churn despite being premium. Investigate quality gap.",span=7)
r=1; hrow(ws3,r,7)
for ci,h in enumerate(["InternetKey","ServiceName","ChurnRate","AvgMonthly","CustomerCount","RiskLevel","KeyInsight"],1): ws3.cell(row=r,column=ci,value=h)
r+=1
rlm={"HIGH":("DC2626","FEE2E2"),"MEDIUM":("92400E","FEF9C3"),"LOW":("059669","D1FAE5")}
for key,risk,insight in [("Fiber optic","HIGH","Premium — highest churn, investigate quality gap"),
("DSL","MEDIUM","Stable price-sensitive segment"),("No internet","LOW","Most loyal — upsell opportunity")]:
    fms=[key,key,
    f'=COUNTIFS(fact_Customers!I:I,A{r},fact_Customers!U:U,"Yes")/COUNTIF(fact_Customers!I:I,A{r})',
    f'=AVERAGEIF(fact_Customers!I:I,A{r},fact_Customers!S:S)',
    f'=COUNTIF(fact_Customers!I:I,A{r})',risk,insight]
    for ci,v in enumerate(fms,1):
        c=dc(ws3,r,ci,v)
        if ci==3: c.number_format="0.0%"; c.alignment=CTR()
        if ci==4: c.number_format='"$"#,##0.00'; c.alignment=RGT()
        if ci==6:
            tc,bc=rlm.get(v,("374151","F9FAFB"))
            c.fill=FILL(bc); c.font=F(bold=True,sz=10,color=tc); c.alignment=CTR()
    ws3.row_dimensions[r].height=18; r+=1
setw(ws3,[14,22,12,14,14,12,46]); ws3.freeze_panes="A4"; print("✓ dim_Internet")

# ═════════════════════════════════════════════════════════
# dim_Payment
# ═════════════════════════════════════════════════════════
ws4=wb.create_sheet("dim_Payment")
# title(ws4,"dim_Payment | Key: PaymentKey | JOIN → fact_Customers.PaymentMethod",span=6)
# note(ws4,"Auto-pay methods (Bank transfer, Credit card) show significantly lower churn rates",span=6)
r=1; hrow(ws4,r,6)
for ci,h in enumerate(["PaymentKey","PaymentName","ChurnRate","CustomerCount","IsAutoPay","RecommendedAction"],1): ws4.cell(row=r,column=ci,value=h)
r+=1
for key,name,auto,action in [("Electronic check","Electronic check","No","Nudge to auto-pay — 5% discount"),
("Mailed check","Mailed check","No","Offer paperless + auto-pay bundle"),
("Bank transfer","Bank transfer (automatic)","Yes","Maintain — lowest friction"),
("Credit card","Credit card (automatic)","Yes","Maintain — lowest friction")]:
    fms=[key,name,
    f'=COUNTIFS(fact_Customers!R:R,A{r},fact_Customers!U:U,"Yes")/COUNTIF(fact_Customers!R:R,A{r})',
    f'=COUNTIF(fact_Customers!R:R,A{r})',auto,action]
    for ci,v in enumerate(fms,1):
        c=dc(ws4,r,ci,v)
        if ci==3: c.number_format="0.0%"; c.alignment=CTR()
        if ci==5:
            c.alignment=CTR()
            if v=="No": c.fill=FILL("FEE2E2"); c.font=F(bold=True,sz=10,color="DC2626")
            else:       c.fill=FILL("D1FAE5"); c.font=F(bold=True,sz=10,color="059669")
    ws4.row_dimensions[r].height=18; r+=1
setw(ws4,[18,26,12,14,12,40]); ws4.freeze_panes="A4"; print("✓ dim_Payment")

# ═════════════════════════════════════════════════════════
# dim_RiskLevel
# ═════════════════════════════════════════════════════════
ws5=wb.create_sheet("dim_RiskLevel")
# title(ws5,"dim_RiskLevel | Key: RiskKey | JOIN → fact_Customers.RiskLevel",span=7)
# note(ws5,"Thresholds: High > 60% | Medium 35-60% | Low < 35%",span=7)
r=1; hrow(ws5,r,7)
for ci,h in enumerate(["RiskKey","RiskLabel","ThresholdFrom","ThresholdTo","CustomerCount","AvgChurnProb","ImmediateAction"],1): ws5.cell(row=r,column=ci,value=h)
r+=1
rm2={"High":("DC2626","FEE2E2"),"Medium":("92400E","FEF9C3"),"Low":("059669","D1FAE5")}
for key,lo,hi,action in [("High",0.60,0.93,"Personal call within 48h + exclusive offer"),
("Medium",0.35,0.60,"Targeted email + discount within 1 week"),
("Low",0.03,0.35,"Loyalty rewards — quarterly check-in")]:
    tc,bc=rm2[key]
    fms=[key,f"{key.upper()} RISK",lo,hi,
    f'=COUNTIF(fact_Customers!W:W,A{r})',
    f'=AVERAGEIF(fact_Customers!W:W,A{r},fact_Customers!V:V)',action]
    for ci,v in enumerate(fms,1):
        c=dc(ws5,r,ci,v)
        if ci in(1,2): c.fill=FILL(bc); c.font=F(bold=True,sz=10,color=tc)
        if ci in(3,4): c.number_format="0%"; c.alignment=CTR()
        if ci==6: c.number_format="0.0%"; c.alignment=CTR()
    ws5.row_dimensions[r].height=20; r+=1
setw(ws5,[10,14,15,15,14,14,46]); ws5.freeze_panes="A4"; print("✓ dim_RiskLevel")

# ═════════════════════════════════════════════════════════
# dim_TenureGroup
# ═════════════════════════════════════════════════════════
ws6=wb.create_sheet("dim_TenureGroup")
# title(ws6,"dim_TenureGroup | Key: TenureKey | JOIN → fact_Customers.TenureGroup",span=7)
# note(ws6,"0-12 months is the critical retention window — 63% of all churn happens here",span=7)
r=1; hrow(ws6,r,7)
for ci,h in enumerate(["TenureKey","CohortLabel","MonthsFrom","MonthsTo","CustomerCount","ChurnRate","CohortInsight"],1): ws6.cell(row=r,column=ci,value=h)
r+=1
for key,label,lo,hi,insight in [("0-12 mo","New (0-12 months)",0,12,"CRITICAL WINDOW — 63% of all churn. Priority onboarding."),
("13-24 mo","Growing (13-24 months)",13,24,"High risk — loyalty program + contract upgrade."),
("25-48 mo","Established (25-48 months)",25,48,"Stable — focus on upsell."),
("49-72 mo","Loyal (49-72 months)",49,72,"Most loyal — ambassador program.")]:
    fms=[key,label,lo,hi,
    f'=COUNTIF(fact_Customers!X:X,A{r})',
    f'=COUNTIFS(fact_Customers!X:X,A{r},fact_Customers!U:U,"Yes")/COUNTIF(fact_Customers!X:X,A{r})',insight]
    for ci,v in enumerate(fms,1):
        c=dc(ws6,r,ci,v)
        if ci==6: c.number_format="0.0%"; c.alignment=CTR()
        if ci in(3,4): c.alignment=CTR()
    ws6.row_dimensions[r].height=18; r+=1
setw(ws6,[12,22,12,10,14,12,50]); ws6.freeze_panes="A4"; print("✓ dim_TenureGroup")

# ═════════════════════════════════════════════════════════
# dim_ChargesGroup
# ═════════════════════════════════════════════════════════
ws7=wb.create_sheet("dim_ChargesGroup")
# title(ws7,"dim_ChargesGroup | Key: ChargesKey | JOIN → fact_Customers.ChargesGroup",span=7)
# note(ws7,"Higher charges correlate with higher churn — price sensitivity is a key driver",span=7)
r=1; hrow(ws7,r,7)
for ci,h in enumerate(["ChargesKey","BandLabel","RangeFrom","RangeTo","CustomerCount","ChurnRate","AvgTotalRevenue"],1): ws7.cell(row=r,column=ci,value=h)
r+=1
for key,label,lo,hi in [("Low <$35","Low (under $35)",0,35),("Mid $35-65","Mid ($35-$65)",35,65),
("High $65-90","High ($65-$90)",65,90),("Premium >$90","Premium (above $90)",90,200)]:
    fms=[key,label,lo,hi,
    f'=COUNTIF(fact_Customers!Y:Y,A{r})',
    f'=COUNTIFS(fact_Customers!Y:Y,A{r},fact_Customers!U:U,"Yes")/COUNTIF(fact_Customers!Y:Y,A{r})',
    f'=AVERAGEIF(fact_Customers!Y:Y,A{r},fact_Customers!T:T)']
    for ci,v in enumerate(fms,1):
        c=dc(ws7,r,ci,v)
        if ci==6: c.number_format="0.0%"; c.alignment=CTR()
        if ci in(3,4): c.number_format='"$"#,##0'; c.alignment=CTR()
        if ci==7: c.number_format='"$"#,##0.00'; c.alignment=RGT()
    ws7.row_dimensions[r].height=18; r+=1
setw(ws7,[14,20,12,10,14,12,18]); ws7.freeze_panes="A4"; print("✓ dim_ChargesGroup")

# ═════════════════════════════════════════════════════════
# agg_ChurnSummary
# ═════════════════════════════════════════════════════════
ws8=wb.create_sheet("agg_ChurnSummary")
# title(ws8,"agg_ChurnSummary — Live KPIs (auto-calculated from fact_Customers)",span=4)
# note(ws8,"All values are Excel formulas referencing fact_Customers. Press F9 to refresh after data changes.",span=4)
r=1; hrow(ws8,r,4)
for ci,h in enumerate(["KPI_Name","Value","FormulaSource","UsedByDashboard"],1): ws8.cell(row=r,column=ci,value=h)
r+=1
kpis=[("Total_Customers",'=COUNTA(fact_Customers!A:A)-3',"col A count","ALL"),
("Churned_Count",'=COUNTIF(fact_Customers!U:U,"Yes")',"col U","ALL"),
("Retained_Count",'=COUNTIF(fact_Customers!U:U,"No")',"col U","ALL"),
("Churn_Rate","=B5/B4","Churned/Total","ALL"),
("Avg_Monthly_All","=AVERAGE(fact_Customers!S2:S2001)","col S","P1-Overview"),
("Avg_Monthly_Churned",'=AVERAGEIF(fact_Customers!U:U,"Yes",fact_Customers!S:S)',"col S","P1-Overview"),
("Avg_Monthly_Retained",'=AVERAGEIF(fact_Customers!U:U,"No",fact_Customers!S:S)',"col S","P1-Overview"),
("Avg_Tenure_All","=AVERAGE(fact_Customers!F2:F2001)","col F","P3-Segments"),
("Avg_Tenure_Churned",'=AVERAGEIF(fact_Customers!U:U,"Yes",fact_Customers!F:F)',"col F","P3-Segments"),
("Total_Revenue","=SUM(fact_Customers!T2:T2001)","col T","P4-Financial"),
("Lost_Revenue",'=SUMIF(fact_Customers!U:U,"Yes",fact_Customers!T:T)',"col T","P4-Financial"),
("Revenue_Retained","=B14-B15","Total-Lost","P4-Financial"),
("Lost_Revenue_Pct","=B15/B14","Lost/Total","P4-Financial"),
("High_Risk_Count",'=COUNTIF(fact_Customers!W:W,"High")',"col W","P5-Predictions"),
("Medium_Risk_Count",'=COUNTIF(fact_Customers!W:W,"Medium")',"col W","P5-Predictions"),
("Low_Risk_Count",'=COUNTIF(fact_Customers!W:W,"Low")',"col W","P5-Predictions"),
("Avg_ChurnProbability","=AVERAGE(fact_Customers!V2:V2001)","col V","P5-Predictions")]
pct={"Churn_Rate","Lost_Revenue_Pct","Avg_ChurnProbability"}
money={"Avg_Monthly_All","Avg_Monthly_Churned","Avg_Monthly_Retained","Total_Revenue","Lost_Revenue","Revenue_Retained"}
for kn,kf,ks,ku in kpis:
    c1=dc(ws8,r,1,kn); c1.font=F(bold=True,sz=10)
    c2=dc(ws8,r,2,kf,ac="center"); c2.font=F(bold=True,sz=11,color="1E40AF"); c2.fill=FILL("EFF6FF")
    if kn in pct: c2.number_format="0.0%"
    elif kn in money: c2.number_format='"$"#,##0.00'
    dc(ws8,r,3,ks).font=F(sz=9,italic=True,color="6B7280")
    dc(ws8,r,4,ku,ac="center").font=F(sz=10,color="7C3AED")
    ws8.row_dimensions[r].height=18; r+=1
setw(ws8,[26,20,32,16]); ws8.freeze_panes="A4"; print("✓ agg_ChurnSummary")

# ═════════════════════════════════════════════════════════
# agg_Financial
# ═════════════════════════════════════════════════════════
ws9=wb.create_sheet("agg_Financial")
# title(ws9,"agg_Financial — Revenue Impact & Retention ROI Scenarios",span=7)
# note(ws9,"BLUE cells = editable inputs. All calculations auto-update when you change blue values.",span=7)
r=1
ws9.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
ws9.cell(row=r,column=1,value="INPUTS — Blue = editable").font=F(bold=True,sz=10,color="FFFFFF")
ws9.cell(row=r,column=1).fill=FILL("1E3A5F"); ws9.cell(row=r,column=1).alignment=LFT()
ws9.row_dimensions[r].height=22; r+=1
for lbl,val,fmt,desc in [("cost_per_customer",150,'"$"#,##0',"Campaign cost per at-risk customer ($)"),
("lifetime_months",24,'0',"Expected months a retained customer stays"),
("rate_conservative",0.10,'0%',"Conservative retention rate"),
("rate_moderate",0.20,'0%',"Moderate retention rate"),
("rate_optimistic",0.30,'0%',"Optimistic retention rate"),
("rate_best_case",0.45,'0%',"Best case retention rate")]:
    c1=ws9.cell(row=r,column=1,value=lbl); c1.font=F(bold=True,sz=10)
    c1.border=BDR(); c1.fill=FILL("F9FAFB") if r%2==0 else FILL("FFFFFF"); c1.alignment=LFT()
    c2=ws9.cell(row=r,column=2,value=val)
    c2.font=Font(name="Arial",bold=True,size=11,color="0000FF")
    c2.alignment=CTR(); c2.border=BDR(); c2.fill=FILL("EFF6FF"); c2.number_format=fmt
    c3=ws9.cell(row=r,column=3,value=desc)
    c3.font=F(sz=9,italic=True,color="6B7280"); c3.border=BDR()
    c3.fill=FILL("F9FAFB") if r%2==0 else FILL("FFFFFF"); c3.alignment=LFT()
    ws9.merge_cells(start_row=r,start_column=3,end_row=r,end_column=7)
    ws9.row_dimensions[r].height=18; r+=1
r+=1
ws9.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
ws9.cell(row=r,column=1,value="ROI SCENARIOS — auto-calculated").font=F(bold=True,sz=10,color="FFFFFF")
ws9.cell(row=r,column=1).fill=FILL("1E3A5F"); ws9.cell(row=r,column=1).alignment=LFT()
ws9.row_dimensions[r].height=22; r+=1
hrow(ws9,r,7)
for ci,h in enumerate(["Scenario","RetentionRate","ChurnedBase","CustomersSaved","RevenueRecovered","CampaignCost","NetROI"],1): ws9.cell(row=r,column=ci,value=h)
r+=1
cf='=agg_ChurnSummary!B5'; amf='=agg_ChurnSummary!B6'
for scen,rref in [("Conservative","=agg_Financial!B7"),("Moderate","=agg_Financial!B8"),
("Optimistic","=agg_Financial!B9"),("Best Case","=agg_Financial!B10")]:
    for ci,v in enumerate([scen,rref,cf,f'=ROUND({cf}*{rref},0)',
    f'=D{r}*{amf}*agg_Financial!B5',f'=D{r}*agg_Financial!B4',f'=E{r}-F{r}'],1):
        c=dc(ws9,r,ci,v)
        if ci==2: c.number_format="0%"; c.alignment=CTR()
        if ci in(5,6,7): c.number_format='"$"#,##0'
        if ci==7: c.font=F(bold=True,sz=10,color="059669"); c.fill=FILL("D1FAE5")
        if ci==6: c.font=F(bold=True,sz=10,color="DC2626")
    ws9.row_dimensions[r].height=18; r+=1
setw(ws9,[20,14,14,16,20,18,18]); ws9.freeze_panes="A4"; print("✓ agg_Financial")

wb.save("churn_data_model.xlsx")
print("\n✅ DONE! Sheets:", wb.sheetnames)